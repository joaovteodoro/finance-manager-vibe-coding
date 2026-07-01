"""
Geração de relatórios Excel formatados para DRE e Balanço Patrimonial.

Utiliza openpyxl para criar arquivos .xlsx com formatação profissional,
cores, totais em negrito e aba separada de índices com benchmarks.
"""

from pathlib import Path
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# Paleta de cores
COR_HEADER = "1e1e2e"
COR_RECEITA = "00c896"
COR_DESPESA = "ff4f6d"
COR_NEUTRO = "2a2a3e"
COR_TEXTO_CLARO = "FFFFFF"
COR_AMARELO = "f59e0b"


def exportar_dre(dados: dict, pasta_destino: str = ".") -> str:
    """Gera o arquivo Excel do DRE.

    Args:
        dados: Dicionário retornado por dre_service.calcular_dre().
        pasta_destino: Caminho da pasta onde o arquivo será salvo.

    Returns:
        Caminho completo do arquivo gerado.
    """
    wb = Workbook()
    _criar_aba_dre(wb, dados)
    _criar_aba_indices(wb, dados["indices"], "DRE")
    wb.remove(wb["Sheet"])

    nome = f"DRE_{dados['mes']:02d}{dados['ano']}.xlsx"
    caminho = Path(pasta_destino) / nome
    wb.save(caminho)
    return str(caminho)


def exportar_balanco(dados: dict, pasta_destino: str = ".") -> str:
    """Gera o arquivo Excel do Balanço Patrimonial.

    Args:
        dados: Dicionário retornado por balanco_service.calcular_balanco().
        pasta_destino: Caminho da pasta onde o arquivo será salvo.

    Returns:
        Caminho completo do arquivo gerado.
    """
    wb = Workbook()
    _criar_aba_balanco(wb, dados)
    _criar_aba_indices(wb, dados["indices"], "Balanço")
    wb.remove(wb["Sheet"])

    nome = f"Balanco_{dados['mes']:02d}{dados['ano']}.xlsx"
    caminho = Path(pasta_destino) / nome
    wb.save(caminho)
    return str(caminho)


def _criar_aba_dre(wb: Workbook, dados: dict) -> None:
    """Preenche a aba principal do DRE.

    Args:
        wb: Workbook openpyxl.
        dados: Dados calculados do DRE.
    """
    ws = wb.create_sheet("DRE")
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 20

    _celula(ws, 1, 1, f"DRE — {dados['mes']:02d}/{dados['ano']}", header=True, colspan=2)
    ws.merge_cells("A1:B1")

    linha = 3
    _celula(ws, linha, 1, "Receita Bruta", negrito=True)
    _celula(ws, linha, 2, dados["receita_bruta"], moeda=True, cor_fundo=COR_RECEITA)
    linha += 1

    _celula(ws, linha, 1, "(-) Despesas Operacionais", negrito=True, italico=True)
    linha += 1
    for nome_cat, val in dados["despesas_operacionais"].items():
        _celula(ws, linha, 1, f"    {nome_cat}")
        _celula(ws, linha, 2, -val, moeda=True)
        linha += 1

    _celula(ws, linha, 1, "Resultado Operacional", negrito=True)
    cor = COR_RECEITA if dados["resultado_operacional"] >= 0 else COR_DESPESA
    _celula(ws, linha, 2, dados["resultado_operacional"], moeda=True, negrito=True, cor_fundo=cor)
    linha += 2

    if dados["despesas_nao_operacionais"]:
        _celula(ws, linha, 1, "(-) Despesas Não Operacionais (extraordinárias)", negrito=True, italico=True)
        linha += 1
        for nome_cat, val in dados["despesas_nao_operacionais"].items():
            _celula(ws, linha, 1, f"    {nome_cat}")
            _celula(ws, linha, 2, -val, moeda=True)
            linha += 1
        linha += 1

    _celula(ws, linha, 1, "Resultado Líquido", negrito=True)
    cor = COR_RECEITA if dados["resultado_liquido"] >= 0 else COR_DESPESA
    _celula(ws, linha, 2, dados["resultado_liquido"], moeda=True, negrito=True, cor_fundo=cor)


def _criar_aba_balanco(wb: Workbook, dados: dict) -> None:
    """Preenche a aba principal do Balanço Patrimonial.

    Args:
        wb: Workbook openpyxl.
        dados: Dados calculados do Balanço.
    """
    ws = wb.create_sheet("Balanço Patrimonial")
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 20

    _celula(ws, 1, 1, f"Balanço Patrimonial — {dados['mes']:02d}/{dados['ano']}", header=True, colspan=2)
    ws.merge_cells("A1:B1")

    linha = 3
    _celula(ws, linha, 1, "ATIVO", negrito=True, cor_fundo=COR_NEUTRO)
    _celula(ws, linha, 2, "", cor_fundo=COR_NEUTRO)
    linha += 1
    _celula(ws, linha, 1, "Ativo Circulante — Caixa")
    _celula(ws, linha, 2, dados["caixa"], moeda=True)
    linha += 1
    if dados["itens_ativo_nao_circulante"]:
        _celula(ws, linha, 1, "Ativo Não Circulante:", italico=True)
        linha += 1
        for nome_item, val in dados["itens_ativo_nao_circulante"].items():
            _celula(ws, linha, 1, f"    {nome_item}")
            _celula(ws, linha, 2, val, moeda=True)
            linha += 1
    _celula(ws, linha, 1, "Total do Ativo", negrito=True)
    _celula(ws, linha, 2, dados["ativo_total"], moeda=True, negrito=True, cor_fundo=COR_RECEITA)
    linha += 2

    _celula(ws, linha, 1, "PASSIVO + PATRIMÔNIO LÍQUIDO", negrito=True, cor_fundo=COR_NEUTRO)
    _celula(ws, linha, 2, "", cor_fundo=COR_NEUTRO)
    linha += 1
    if dados["itens_passivo"]:
        _celula(ws, linha, 1, "Dívidas:", italico=True)
        linha += 1
        for nome_item, val in dados["itens_passivo"].items():
            _celula(ws, linha, 1, f"    {nome_item}")
            _celula(ws, linha, 2, val, moeda=True)
            linha += 1
    else:
        _celula(ws, linha, 1, "Dívidas")
        _celula(ws, linha, 2, 0.0, moeda=True)
        linha += 1
    _celula(ws, linha, 1, "Patrimônio Líquido")
    _celula(ws, linha, 2, dados["patrimonio_liquido"], moeda=True)
    linha += 1
    _celula(ws, linha, 1, "Total Passivo + PL", negrito=True)
    total_passivo_pl = dados["dividas"] + dados["patrimonio_liquido"]
    _celula(ws, linha, 2, total_passivo_pl, moeda=True, negrito=True, cor_fundo=COR_RECEITA)


def _criar_aba_indices(wb: Workbook, indices: list, tipo: str) -> None:
    """Cria a aba 'Painel de Índices' no Excel.

    Args:
        wb: Workbook openpyxl.
        indices: Lista de índices calculados pelo service.
        tipo: 'DRE' ou 'Balanço' para o título.
    """
    ws = wb.create_sheet("Painel de Índices")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 50

    _celula(ws, 1, 1, f"Painel de Índices — {tipo}", header=True, colspan=5)
    ws.merge_cells("A1:E1")

    headers = ["Índice", "Valor", "Benchmark", "Status", "Interpretação"]
    for col, h in enumerate(headers, 1):
        _celula(ws, 3, col, h, negrito=True, cor_fundo=COR_NEUTRO)

    for i, idx in enumerate(indices, 4):
        _celula(ws, i, 1, idx["nome"])
        if idx["valor"] is None:
            _celula(ws, i, 2, "N/A — dado não informado")
        else:
            valor_fmt = f"{idx['valor']:.2f}{idx['unidade']}"
            _celula(ws, i, 2, valor_fmt)
        _celula(ws, i, 3, idx.get("benchmark", ""))
        _celula(ws, i, 4, idx["icone"])
        _celula(ws, i, 5, idx["interpretacao"])
        ws[f"E{i}"].alignment = Alignment(wrap_text=True)


def _celula(
    ws,
    linha: int,
    col: int,
    valor,
    negrito: bool = False,
    italico: bool = False,
    moeda: bool = False,
    cor_fundo: str = None,
    header: bool = False,
    colspan: int = 1,
) -> None:
    """Escreve e formata uma célula no worksheet.

    Args:
        ws: Worksheet openpyxl.
        linha: Número da linha (1-indexado).
        col: Número da coluna (1-indexado).
        valor: Valor a ser escrito.
        negrito: Aplica negrito ao texto.
        italico: Aplica itálico ao texto.
        moeda: Formata como moeda brasileira.
        cor_fundo: Cor hexadecimal de fundo (sem #).
        header: Aplica estilo de cabeçalho.
        colspan: Quantidade de colunas para mesclar.
    """
    cell = ws.cell(row=linha, column=col)

    if moeda and isinstance(valor, (int, float)):
        cell.value = valor
        cell.number_format = 'R# ##0.00'
    else:
        cell.value = valor

    font_color = COR_TEXTO_CLARO if (cor_fundo or header) else "000000"
    cell.font = Font(
        bold=negrito or header,
        italic=italico,
        color=font_color,
        size=12 if header else 11,
    )

    if header:
        cell.fill = PatternFill("solid", fgColor=COR_HEADER)
    elif cor_fundo:
        cell.fill = PatternFill("solid", fgColor=cor_fundo)

    cell.alignment = Alignment(horizontal="left", vertical="center")
